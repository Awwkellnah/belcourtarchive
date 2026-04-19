#!/usr/bin/env python3
"""
belcourt_extract.py
-------------------
Extracts film titles with opening/closing dates from the Belcourt Theatre
"Gross Revenue by Event" Excel report.

Usage:
    python3 belcourt_extract.py <path_to_xlsx> [--output <output.csv>]

Options:
    --output   Output CSV filename (default: belcourt_films_<date>.csv)
    --all      Include flagged non-film entries in output (marked with a
               FLAGGED column) instead of writing them to a separate file

Output:
    Two files are written:
      1. belcourt_films_<date>.csv      — clean film entries
      2. belcourt_flagged_<date>.csv    — entries that may not be films,
                                          for your manual review

Requirements:
    pip install openpyxl
"""

import sys
import csv
import re
import argparse
from datetime import datetime, date
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Missing dependency: run  pip install openpyxl  then try again.")


# ---------------------------------------------------------------------------
# Patterns that suggest an entry is NOT a standalone film
# ---------------------------------------------------------------------------

# Exact or prefix matches (case-insensitive) that are almost certainly events
NON_FILM_PREFIXES = [
    "belcourt benefit",
    "belcourt talks",
    "belcourt membership",
    "rental",
    "private rental",
    "private event",
    "seminar",
    "workshop",
    "q&a",
    "panel",
    "reception",
    "fundraiser",
    "volunteer",
    "staff ",
    "meeting",
    "test event",
]

# Regex patterns applied to the full title
NON_FILM_PATTERNS = [
    r"\brelease party\b",          # "Taylor Swift | The Official Release Party"
    r"\brelease show\b",
    r"^catvideoFest",              # annual cat video compilation
    r"^\d+ hours of terror$",      # marathon events (multi-film)
    r"^double feature:",           # double feature events
    r"^triple feature:",           # triple feature events
    r"^roger corman double",
    r"^don hertzfeldt presents",
    r"animation mixtape",
    r"short film tour",
    r"^taylor swift",
    r"\bfilm festival\b",
    r"short films?:",              # "Oscar Short Films: ANIMATION"
    r"^thom yorke",                # concerts/performances
    r"nashville famous",           # stand-up comedy
    r"last class with",            # lecture
    r"opryland usa",               # documentary/local special
    r"the last class",
    r"heightened scrutiny",        # legal event
]

# Compile patterns once
_NON_FILM_RE = [re.compile(p, re.IGNORECASE) for p in NON_FILM_PATTERNS]


def looks_like_non_film(title: str) -> bool:
    """Return True if the title looks like a non-film event."""
    lower = title.lower()
    for prefix in NON_FILM_PREFIXES:
        if lower.startswith(prefix):
            return True
    for pattern in _NON_FILM_RE:
        if pattern.search(title):
            return True
    return False


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def is_title_row(row: tuple) -> bool:
    """
    A film title row has a non-empty string in column A and all other
    columns empty (or absent).  Total rows end with ' Total'.
    """
    if not row or not row[0]:
        return False
    val = str(row[0]).strip()
    if not val:
        return False
    # Skip header / report metadata rows
    skip_starts = (
        "sale date range",
        "belcourt theatre",
        "gross revenue",
        "summary by day",
        "film:",
        "event date",
    )
    if any(val.lower().startswith(s) for s in skip_starts):
        return False
    # Skip totals rows
    if val.endswith(" Total") or val == "Total":
        return False
    # Must be a single-column row (all other cells None or empty string)
    rest = row[1:]
    if not all(v is None or v == "" for v in rest):
        return False
    return True


def is_date_row(row: tuple) -> bool:
    """A screening row has a date in column A."""
    if not row or not row[0]:
        return False
    val = row[0]
    if isinstance(val, (datetime, date)):
        return True
    if isinstance(val, str):
        # Try common date formats
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"):
            try:
                datetime.strptime(val.strip(), fmt)
                return True
            except ValueError:
                pass
    return False


def parse_date(val) -> date:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                pass
    raise ValueError(f"Cannot parse date: {val!r}")


def extract_films(xlsx_path: str) -> list[dict]:
    """
    Parse the workbook and return a list of dicts with keys:
        title, opening_date, closing_date, flagged, flag_reason
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    films = []
    current_title = None
    current_dates = []

    def flush():
        if current_title and current_dates:
            opening = min(current_dates)
            closing = max(current_dates)
            flagged = looks_like_non_film(current_title)
            films.append({
                "title": current_title,
                "opening_date": opening.strftime("%Y-%m-%d"),
                "closing_date": closing.strftime("%Y-%m-%d"),
                "flagged": "YES" if flagged else "",
                "flag_reason": "Possible non-film event — review manually" if flagged else "",
            })

    for row in ws.iter_rows(values_only=True):
        if is_title_row(row):
            flush()
            current_title = str(row[0]).strip()
            current_dates = []
        elif is_date_row(row):
            try:
                current_dates.append(parse_date(row[0]))
            except ValueError:
                pass  # skip unparseable rows

    flush()  # don't forget the last film
    wb.close()

    # Sort by opening date, then title
    films.sort(key=lambda f: (f["opening_date"], f["title"].upper()))
    return films


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Extract film titles + dates from a Belcourt Gross Revenue xlsx."
    )
    parser.add_argument("xlsx", help="Path to the .xlsx file")
    parser.add_argument(
        "--output",
        help="Output CSV filename for clean films (default: belcourt_films_YYYYMMDD.csv)",
        default=None,
    )
    parser.add_argument(
        "--all",
        action="store_true",
        help="Write ALL entries (including flagged) to a single CSV with a FLAGGED column",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        sys.exit(f"File not found: {xlsx_path}")

    print(f"Reading {xlsx_path.name} …")
    films = extract_films(str(xlsx_path))

    today = datetime.today().strftime("%Y%m%d")
    output_stem = args.output or f"belcourt_films_{today}"
    output_stem = output_stem.removesuffix(".csv")

    clean = [f for f in films if not f["flagged"]]
    flagged = [f for f in films if f["flagged"]]

    fieldnames = ["title", "opening_date", "closing_date"]

    if args.all:
        # Single file with FLAGGED column
        out_path = Path(f"{output_stem}.csv")
        all_fieldnames = fieldnames + ["flagged", "flag_reason"]
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=all_fieldnames)
            writer.writeheader()
            writer.writerows(films)
        print(f"\n✓ All {len(films)} entries written to:  {out_path}")
        print(f"  ({len(clean)} clean,  {len(flagged)} flagged for review)")

    else:
        # Two separate files
        clean_path = Path(f"{output_stem}.csv")
        flagged_path = Path(f"{output_stem}_flagged.csv")

        with open(clean_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for row in clean:
                writer.writerow({k: row[k] for k in fieldnames})

        with open(flagged_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames + ["flag_reason"])
            writer.writeheader()
            for row in flagged:
                writer.writerow({k: row[k] for k in fieldnames + ["flag_reason"]})

        print(f"\n✓ {len(clean)} clean films  →  {clean_path}")
        print(f"  {len(flagged)} flagged entries  →  {flagged_path}")
        print(f"\nReview {flagged_path.name}, then move any real films into {clean_path.name}.")

    # Print a quick summary preview
    print(f"\n--- Preview (first 10 clean films) ---")
    for film in clean[:10]:
        print(f"  {film['opening_date']}  {film['title']}")
    if len(clean) > 10:
        print(f"  … and {len(clean) - 10} more")


if __name__ == "__main__":
    main()